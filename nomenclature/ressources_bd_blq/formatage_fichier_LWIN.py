from unidecode import unidecode

import openpyxl
from openpyxl import Workbook

import pandas as pd

from datetime import datetime

nomProfil : str = 'LWINdatabase'
# nomProfil : str = 'test_base_LWIN'
extensionFin_bis : str = '.xlsx'
extensionFin : str = '.csv'

# Rajout d'une partie du chemin car python ce positionne dans le dossier du premier fichier
chemin : str = "ressources_bd_blq/"
# Pour les tests
# chemin : str = ""

extensionDebut : str = '.xlsx' 

production_str : str = "_production"

dest_filename_production_debut : str = chemin + nomProfil + production_str + extensionDebut
dest_filename_production_bis : str = chemin + 'sortie_bis_' + nomProfil + production_str + extensionFin_bis
dest_filename_production : str = chemin + 'sortie_' + nomProfil + production_str + extensionFin
dest_filename_production_final : str = chemin + 'sortie_final_' + production_str + nomProfil + extensionFin

tab_chemin_production : list = [dest_filename_production_debut, dest_filename_production_bis, dest_filename_production, dest_filename_production_final]

test_str : str = "_test"

dest_filename_test_debut : str = chemin + nomProfil + test_str + extensionDebut
dest_filename_test_bis : str = chemin + 'sortie_bis_' + nomProfil + test_str + extensionFin_bis
dest_filename_test : str = chemin + 'sortie_' + nomProfil + test_str + extensionFin
dest_filename_test_final : str = chemin + 'sortie_final_' + nomProfil + test_str  + extensionFin

tab_chemin_test : list = [dest_filename_test_debut, dest_filename_test_bis, dest_filename_test, dest_filename_test_final]

# Un Warning sur le workbook peut apparaître, il n'est pas important : Workbook contains no default style

def formaterFichierLWIN():

    # print(tab_chemin_production)

    formaterfichier(tab_chemin_production)
    formaterfichier(tab_chemin_test)

def formaterfichier(tab_chemin : list):

    print("Démarrage du formattage du fichier de l'ensemble des vins LWIN : " + str(datetime.now()))

    [dest_filename_debut, dest_filename_bis, dest_filename, dest_filename_final] = tab_chemin
    
    # coding: utf-8
    wb = openpyxl.load_workbook( dest_filename_debut )

    # On prépare un nouveau workbook
    wb2 = Workbook()
    ws = wb2.active
    ws.title = nomProfil

    # Lecture du workbook d'origine
    sheet_list = wb.sheetnames  # on récupère le nom des onglets, pas necessaire ici.
    sheet = wb[sheet_list[0]]

    iVinAppellation : tuple = sheet['C']
    iCouleur : tuple = sheet['L']
    iProducteur : tuple = sheet['E']
    iRegion : tuple = sheet['H']
    region : str = ''
    iPays : tuple = sheet['G']

    iProducteurTitre : tuple = sheet['D']

    ###
    iAppellation : tuple = sheet['I']
    iAppellationComplement : tuple = sheet['J']
    ###

    row_count = sheet.max_row
    column_count = sheet.max_column

    index_sortie : int = 1

    for i  in range (1, row_count):
                
        vin : str = unidecode(str(iVinAppellation[i].value)).upper()

        print(vin)

        region = unidecode(str(iRegion[i].value)).upper()
        couleur : str = unidecode(str(iCouleur[i].value)).upper()

        if( vin == "DISPLAY_NAME" or couleur == "NA"):
            continue;
        
        # if( vin == "DISPLAY_NAME" or region != "BORDEAUX" or couleur == "NA"):
        #     continue;

        producteur : str = unidecode(str(iProducteur[i].value)).upper()
        producteur_titre : str = unidecode(str(iProducteurTitre[i].value)).upper()

        if(producteur_titre != "NA") : 
            producteur = producteur_titre + " " + producteur

        pays : str = unidecode(str(iPays[i].value)).upper()

        appellation : str = unidecode(str(iAppellation[i].value)).upper()
        appellationComplement : str = unidecode(str(iAppellationComplement[i].value).strip()).upper()

        if( (len(appellationComplement) != 0) & (appellationComplement != "NA") ) : 
            appellation = appellation + " " + appellationComplement

        # print("  vin : " + vin + " couleur " + couleur + " producteur " + producteur + " pays " + pays + " appellation " + appellation + "")

        # appellation
        # Fonction Attribution des valeurs dans les bonnes colonnes pour le nouveau workbook
        c1 = ws.cell(row = index_sortie, column = 1)
        c1.value = vin

        c2 = ws.cell(row = index_sortie, column = 2)
        c2.value = couleur

        c3 = ws.cell(row = index_sortie, column = 3)
        c3.value = producteur

        c4 = ws.cell(row = index_sortie, column = 4)
        c4.value = region

        c5 = ws.cell(row = index_sortie, column = 5)
        c5.value = pays

        ###
        c6 = ws.cell(row = index_sortie, column = 6)
        c6.value = appellation
        ###

        index_sortie += 1

    # Fonction de suppression des lignes vides d'un workbook
    # index_row = []
    # for n in range(1, ws.max_row):
    #     if ws.cell(n, 1).value is None:
    #         index_row.append(n)

    # for row_del in range(len(index_row)):
    #     ws.delete_rows(idx=index_row[row_del], amount=1)
    #     index_row = list(map(lambda k: k -1, index_row))

        # Ecriture du nouveau workbook
        wb2.save(dest_filename_bis)

    read_file = pd.read_excel(dest_filename_bis)
    read_file.to_csv(dest_filename, index = False, encoding="utf-8", sep = ';')

    print("Fin du formattage du fichier de l'ensemble des vins LWIN : " + str(datetime.now()))

# Le décommenter que pour les tests
# Enlever la variable "chemin" dans les chemins d'accès aux fichiers
# formaterFichierLWIN()