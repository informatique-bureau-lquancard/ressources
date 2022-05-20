import glob
from unidecode import unidecode
import openpyxl
from openpyxl import Workbook

### Non utilisé

import sys
sys.path.append("/var/www/html/projets/bureau-laurent-quancard/gestion-des-offres/gestion_tarifs_automatises/Tarifs/batchs")
import requetes_blq as req

sys.path.append("/var/www/html/projets/bureau-laurent-quancard/gestion-des-offres/gestion_tarifs_automatises/Tarifs/batchs")
import nomenclature_bd_blq as nom

sys.path.append("/var/www/html/projets/bureau-laurent-quancard/gestion-des-offres/gestion_tarifs_automatises/Tarifs/batchs")
import requetes_blq as req

# Pour rajouter la bibliothèque numpy : python3 -m pip install numpy
# Permet d'inverser les tableaux
import numpy as np

import peuplement_de_la_base as pp

tab_appellation : list = req.recuperation_tab( nom.appellation_str )
tab_appellation_inverse : list = list(map(list, np.transpose(tab_appellation)))

# Même fonction que traitement_vins : trouver un emplacement
def recupId(tab_inverse : list, valeur_testee, nom_table : str, nom_colonne : str, id_colonne : int):
    id : int = pp.recup_val_tab_inv2(tab_inverse, id_colonne, valeur_testee)

    print("appellation : " + valeur_testee)

    if(id == -1):

        # La "," est importante car sinon cela ne marche pas
        tuple_valeurs : tuple = (valeur_testee,)
        requete = f"""INSERT INTO """+ nom_table +""" ("""+ nom_colonne +""") VALUES (%s);"""
        
        req.envoie_requete_tuple_sans_retour(requete, tuple_valeurs)

        id : int = pp.recup_val_tab_inv2(tab_inverse, id_colonne, valeur_testee)
        print("Retour id : " + str(id))

    return id

def insertionAppellations():
    
    extensionDebut : str = '.xlsx'

    # coding: utf-8
    # On load le le tarif au format xlsx
    for filename in glob.glob("*" + extensionDebut):
        wb = openpyxl.load_workbook(filename)

    # Lecture du workbook d'origine
    sheet_list = wb.sheetnames  # on récupère le nom des onglets, pas necessaire ici.
    sheet = wb[sheet_list[0]]

    iAppellation = sheet['A']

    row_count = sheet.max_row

    print("row_count : " + str(row_count))

    id : int = recupId(tab_appellation_inverse, "NA", nom.appellation_str, "appellation", 1)

    for i  in range (1, row_count):
                
        appellation : str = unidecode(str(iAppellation[i].value)).upper()

        if( (appellation is None) or (appellation == "APPELLATIONS")):
            continue;

        # Recuperation id pour les tests
        id : int = recupId(tab_appellation_inverse, appellation, nom.appellation_str, "appellation", 1)

# Pour le lancer seul
# insertionAppellations()
