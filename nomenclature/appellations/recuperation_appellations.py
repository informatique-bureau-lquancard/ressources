import glob
import openpyxl
from unidecode import unidecode

# import ressources_bd_blq.formatage_fichier_LWIN as ffl

import sys

sys.path.append("/var/www/html/projets/bureau-laurent-quancard/gestion-des-offres/gestion_tarifs_automatises/Tarifs/batchs")
import requetes_blq_test as req

sys.path.append("/var/www/html/ressources/nomenclature/ressources_bd_blq")
import formatage_fichier_LWIN as ffl

sys.path.append("/var/www/html/ressources/nomenclature")
import nomenclature_bd_blq as nom

sys.path.append("/var/www/html/php/fonctions_tarifs")
import Fonction_tarifs as ft

# nom_fichier = "Vins_listes"
nom_fichier = "LWINdatabase"
extension = ".xlsx"

def recuperation_appellations():

    req.reinitialisation_global([nom.appellation_str])
    
    # Ancien fichier servant à récupérer les appellations
    # wb = openpyxl.load_workbook("/var/www/html/ressources/nomenclature/appellations/" + nom_fichier + extension)

    # Fichier LWIN tiré de la base
    wb = openpyxl.load_workbook( ffl.dest_filename_production_bis )
    # wb = openpyxl.load_workbook( "../" + ffl.dest_filename_test_bis )

    # Lecture du workbook d'origine
    sheet_list = wb.sheetnames  # on récupère le nom des onglets, pas necessaire ici.
    sheet = wb[sheet_list[0]]

    iAppellation = sheet['F']

    row_count = sheet.max_row

    # print("row_count : " + str(row_count))

    liste_appellation : list = []

    liste_appellation.append("NA")

    for i  in range (1, row_count):
                
        appellation : str = unidecode(str(iAppellation[i].value)).upper().strip()

        appellation = appellation.replace("'", "")

        if( (appellation is None) | (appellation == "NA") ):
            continue;

        if( ft.rechercheIndiceValeursDansListe(liste_appellation, [appellation] ) != -1 ):
            continue;

        # print( "appellation 3 : " + str( appellation ) )
        
        liste_appellation.append(appellation)

    return liste_appellation

# recuperation_appellations()
        