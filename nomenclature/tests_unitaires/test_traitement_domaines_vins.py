import openpyxl
from openpyxl import Workbook

import sys
sys.path.append("/var/www/html/projets/bureau-laurent-quancard/gestion-des-offres/gestion_tarifs_automatises/Tarifs/batchs")
import requetes_blq_test as req

import nomenclature_bd_blq as nom

sys.path.append("/var/www/html/php/fonctions_tarifs")
import Fonction_tarifs_dev as ft

# Domaines a tester
domaine_d_issan : list = ["CHATEAU D ISSAN", "CHATEAU DISSAN"]

liste_domaine_a_tester : list = [domaine_d_issan]

# Couleurs a tester
liste_rouge : list = ["ROUGE", "RED"]
liste_blanc : list = ["BLANC", "WHITE"]
liste_rose : list = ["ROSE"]
liste_mixed : list = ["MIXED"]
liste_couleur : list = [liste_rouge, liste_blanc, liste_rose, liste_mixed]

liste_appellation_a_enlever = ["CHATEAU D ISSAN", "MOULIS EN MEDOC", "CASTILLON-COTES DE BORDEAUX", "BORDEAUX ROUGE", "LALANDE DE POMEROL", "BLAYE-COTES DE BORDEAUX", "COTES DE BOURG"]

erreur_appellation = ["APPELATIONS EN ERREUR : "]

class TestTraitementDomainesVins:

    def __init__(self):
        print("Création : " + type(self).__name__)

        # Mettre en majuscule
        nom_vin_teste = "RAUZAN-GASSIES"

        self.insertionVins(nom_vin_teste)

    def fct_appellation_id(self, appellation : str):
    
        try:

            print("appellation : " + appellation)

            if( appellation in liste_appellation_a_enlever ):
                appellation = "MARGAUX"

            print("appellation : " + appellation)

            appellation_id = req.recuperation_un_id(nom.appellation_str, "appellation", appellation)

            print("appellation_id : " + str(appellation_id))

            if( appellation_id == -1 ):
                erreur_appellation.append("Appellation en erreur : " + appellation)
                appellation_id = 1

            return appellation_id

        except ValueError:
            erreur_appellation.append(appellation + "")

        return 1

    def insertion( self, vin, domaine_id, appellation_id, couleur_id, ws, index_sortie, couleur, iDomaine, i, region, iPays ):
    
        tuple_valeurs : tuple = (vin, "regex", domaine_id, appellation_id, couleur_id)

        requete = f"""INSERT INTO """ + nom.vin_str + """(nom, regex, domaine_id, appellation_id, couleur_id) VALUES (%s, %s, %s, %s, %s);"""
        # req.envoie_requete_tuple_sans_retour(requete, tuple_valeurs) 

        print("vin : "+ vin + "couleur : " + couleur + "iDomaine[i].value : " + iDomaine[i].value + "region : " + region + "iPays[i].value : " + iPays[i].value)

        # Fonction Attribution des valeurs dans les bonnes colonnes pour le nouveau workbook
        c1 = ws.cell(row = index_sortie, column = 1)
        c1.value = vin

        c2 = ws.cell(row = index_sortie, column = 2)
        c2.value = couleur

        c3 = ws.cell(row = index_sortie, column = 3)
        c3.value = iDomaine[i].value

        c4 = ws.cell(row = index_sortie, column = 4)
        c4.value = region

        c5 = ws.cell(row = index_sortie, column = 5)
        c5.value = iPays[i].value

    def insertionVins(self, nom_vin_teste : str):

        # Initialisation
        print()
        # Traitement
        # Test

        nomProfil : str = 'LWINdatabase'
        extensionFin_bis : str = '.xlsx'
        # dest_filename_bis = 'sortie_bis_' + nomProfil + extensionFin_bis
        dest_filename_bis = '/var/www/html/ressources/nomenclature/domainesVins/sortie_bis_LWINdatabase.xlsx'

        # On prépare un nouveau workbook
        ###
        wb3 = Workbook()
        ws = wb3.active
        ws.title = nomProfil
        ###

        # Ouverture du fichier modifier LWIN
        wb_bis = openpyxl.load_workbook(dest_filename_bis)

        print("dest_filename_bis : " + dest_filename_bis)

        # Lecture du workbook d'origine
        sheet_list = wb_bis.sheetnames  # on récupère le nom des onglets, pas necessaire ici.
        sheet = wb_bis[sheet_list[0]]

        iVinAppellation = sheet['A']
        iCouleur = sheet['B']
        iDomaine = sheet['C']
        iRegion = sheet['D']
        region : str = ""
        iPays = sheet['E']

        row_count = sheet.max_row

        index_sortie = 1

        for i in range (1, row_count):

            if(  (iVinAppellation[i].value).find(nom_vin_teste) == -1 ):
                continue
                    # print("iVinAppellation[i].value : " + iVinAppellation[i].value)
            
            vin : str = iVinAppellation[i].value
            list_vinApellation  = vin.split(",")
            vin = vin.replace(",", '')

            if( len(vin) == 0 ):
                vin = "NA"
                return

            region = (iRegion[i].value).upper()

            domaine : list = (iDomaine[i].value)

            print("tab_vinApellation[0] : " + list_vinApellation[0])
            print(domaine)

            # !!! A mettre en place !!!
            # Revoir si exception, les mettre à la main manuellement
            # 
            if( len( list_vinApellation ) == 1):
                print( " 1 " )

            elif( region == "BORDEAUX"):
                vin = list_vinApellation[0]
            
            elif( list_vinApellation[0] == domaine ):

                vin : str = ''.join(list_vinApellation[1:]).strip() 

                print("vin : " + vin)
            else:
                print("Erreur")
                continue
            
            appellation : str = "NA"

            couleur : str = iCouleur[i].value
            pays_id : int = iPays[i].value

            # Insérer les lignes de domaines

            print("domaine  : -" + domaine + "-")

            domaine_id : int = req.recuperation_un_id(nom.domaine_str, "nom", domaine)

            print("domaine_id : " + str(domaine_id))

            if( domaine_id == -1 ):
                domaine = ft.retour_premier_element_avec_valeur( liste_domaine_a_tester, domaine)

                print("domaine : " + domaine)

                domaine_id : int = req.recuperation_un_id(nom.domaine_str, "nom", domaine)

            print("domaine_id : -" + str(domaine_id) + "-")

            # Appellation
            appellation_id = self.fct_appellation_id(appellation)

            print("appellation_id : " + str(appellation_id))

            # Couleur
            couleur_id : int = ft.retour_premier_element_avec_valeur(liste_couleur, couleur)

            print("couleur : " + couleur)

            couleur_id : int = req.recuperation_un_id(nom.couleur_vin_str, "couleur", couleur)
            
            print("couleur_id : " + str(couleur_id))

            if(couleur_id == -1):
                couleur = ft.retour_premier_element_avec_valeur(liste_couleur, couleur)

                print("couleur : -" + couleur + "-")

                couleur_id : int = req.recuperation_un_id(nom.couleur_vin_str, "couleur", couleur)

                if( couleur == "NA" ):
                    couleur_id = 1

                self.insertion( vin, domaine_id, appellation_id, couleur_id, ws, index_sortie, couleur, iDomaine, i, region, iPays )

            print("vin : -" + str(vin) + "-")

            index_sortie += 1

test : TestTraitementDomainesVins = TestTraitementDomainesVins()