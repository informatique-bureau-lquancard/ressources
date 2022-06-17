import csv
import glob
import time
# from subprocess import CREATE_NEW_CONSOLE
import openpyxl
from openpyxl import Workbook

import warnings
warnings.simplefilter("ignore")

import sys
sys.path.append("/var/www/html/projets/bureau-laurent-quancard/gestion-des-offres/gestion_tarifs_automatises/Tarifs/batchs")
import requetes_blq_test as req

import nomenclature_bd_blq as nom

sys.path.append("/var/www/html/php/fonctions_tarifs")
import Fonction_tarifs as ft

# Pour rajouter la bibliothèque numpy : python3 -m pip install numpy
# Permet d'inverser les tableaux
import numpy as np

import peuplement_de_la_base as pp 

from datetime import datetime

import ressources_bd_blq.formatage_fichier_LWIN as ffl

###
#   Les ressources pour les traitements ont été récupérés depuis le site liv-ex.com
###

# dict_couleur : dict = {"NA": 1, "RED": 2, "WHITE": 3, "ROSE": 4}
tab_couleur_vin : list = req.recuperation_tab( nom.couleur_vin_str )
tab_couleur_vin_inverse : list = list(map(list, np.transpose(tab_couleur_vin)))

# dict_pays : dict = {"NA": 1, "FRANCE": 2, "ANGLETERRE": 3, "ALLEMAGNE": 4}
tab_domaine : list = []
tab_domaine_inverse : list = []

# tab_appellation : list = req.recuperation_tab( nom.appellation_str )
# tab_appellation_inverse : list = list(map(list, np.transpose(tab_appellation)))

liste_rouge : list = ["ROUGE", "RED"]
liste_blanc : list = ["BLANC", "WHITE"]
liste_rose : list = ["ROSE"]
liste_mixed : list = ["MIXED"]
liste_couleur : list = [liste_rouge, liste_blanc, liste_rose, liste_mixed]

# Requête pour connaître les doublons
# SELECT nom,COUNT(*) FROM domaine GROUP BY nom HAVING COUNT(*) > 1;

rouge_str : str = "ROUGE"
blanc_str : str = "BLANC"
margaux_str : str = "MARGAUX"
pauillac_str : str = "PAUILLAC"
saint_julien_str : str = "SAINT-JULIEN"
gruaud_larose_str : str = "CHATEAU GRUAUD LAROSE"
leoville_barton_str : str = "CHATEAU LEOVILLE BARTON"
saint_emillion_grand_cru_str : str = "SAINT-EMILION GRAND CRU"
pommerol_str : str = "POMEROL"
d_issan_str : str = "CHATEAU D ISSAN"
rieussec_str : str = "CHATEAU RIEUSSEC"
sauternes_str : str = "SAUTERNES"
chateaux_margaux_str : str = "CHATEAU MARGAUX"
chateau_beychevelle_str : str = "CHATEAU BEYCHEVELLE"

# Domaines rajoutés
liste_domaines_rajoutes : list = ["CHATEAU LASCOMBES", "CHATEAU RAUZAN GASSIES", "SC DU CHATEAU LAFITE ROTHSCHILD", "CHATEAU BRANAIRE DUCRU",
"CHATEAU DUCRU BEAUCAILLOU", "CLOS DE SARPE", "CLOS L'EGLISE", "CHATEAU CANTENAC BROWN", "SCEA BOYD CANTENAC & POUGET", "CHATEAU D ISSAN"]

# Domaine a tester
domaine_d_issan : list = ["CHATEAU D ISSAN", "CHATEAU DISSAN"]

liste_domaine_a_tester : list = [domaine_d_issan]

# Vins rajoutés
vin_chateau_lascombes : list = ["chateau lascombes".upper(), "regex", "NAN", margaux_str, rouge_str]
vin_chateau_rauzan_gassies : list = ["chateau rauzan gassies".upper(), "regex", "NAN", margaux_str, rouge_str]
vin_chateau_lafite_rotchild : list = ["chateau lafite rotchild".upper(), "regex", "NAN", pauillac_str, rouge_str]
vin_chateau_branaire_ducru : list = ["chateau branaire ducru".upper(), "regex", "NAN", saint_julien_str, rouge_str]

# Vin a tester
vin_d_issan : list = ["CHATEAU D ISSAN", "CHATEAU DISSAN"]

liste_vin_a_tester : list = [vin_d_issan]

#a revoir beycheville
vin_chateau_beychevelle : list = ["chateau beychevelle".upper(), "regex", chateau_beychevelle_str, saint_julien_str, rouge_str]

# vin_chateau_saint_pierre = ["CHATEAU RAUZAN-SEGLA", "regex", 651, 319, rouge_str]
vin_chateau_talbot : list = ["chateau talbot".upper(), "regex", "NAN", saint_julien_str, rouge_str]
vin_chateau_lagrange : list = ["chateau lagrange".upper(), "regex", "NAN", saint_julien_str, rouge_str]
vin_chateau_ducru_beau_caillou : list = ["chateau ducru beau caillou".upper(), "regex", "NAN", saint_julien_str, rouge_str]
vin_chateau_gruaud_larose : list = ["vin chateau gruaud larose".upper(), "regex", gruaud_larose_str, saint_julien_str, rouge_str]
vin_chateau_leoville_barton : list = ["chateau leoville barton".upper(), "regex", leoville_barton_str, saint_julien_str, rouge_str]

# vin_chateau_prieure_lichine = ["CHATEAU RAUZAN-SEGLA", "regex", 651, 319, rouge_str]
vin_chateau_faurie_de_souchard : list = ["chateau faurie de souchard".upper(), "regex", "NAN", saint_emillion_grand_cru_str, rouge_str]
vin_chateau_clos_de_sarpe : list = ["chateau clos de sarpe".upper(), "regex", "NAN", saint_emillion_grand_cru_str, rouge_str]
vin_clos_leglise : list = ["clos leglise".upper(), "regex", "NAN", pommerol_str, rouge_str]
vin_chateau_cantenac_brown : list = ["chateau cantenac brown".upper(), "regex", "NAN", margaux_str, rouge_str]
vin_chateau_boyd_cantenac : list = ["chateau boyd cantenac".upper(), "regex", "NAN", margaux_str, rouge_str]
# Prolème décriture : dissan et d'issan 
vin_chateau_d_issan : list = ["chateau dissan".upper(), "regex", "NAN", d_issan_str, rouge_str]

vin_carmes_de_rieussec : list = ["carmes de rieussec".upper(), "regex", rieussec_str, sauternes_str, blanc_str]
vin_rauzan_segla : list = ["CHATEAU RAUZAN-SEGLA", "regex", chateaux_margaux_str, margaux_str, rouge_str]

liste_vins_rajoutes : list = [vin_rauzan_segla, vin_carmes_de_rieussec, vin_chateau_d_issan, vin_chateau_boyd_cantenac, vin_chateau_cantenac_brown, vin_clos_leglise,
vin_chateau_clos_de_sarpe, vin_chateau_faurie_de_souchard, vin_chateau_leoville_barton, vin_chateau_gruaud_larose, vin_chateau_ducru_beau_caillou, 
vin_chateau_lagrange, vin_chateau_talbot, vin_chateau_beychevelle, vin_chateau_branaire_ducru, vin_chateau_lafite_rotchild, vin_chateau_rauzan_gassies, 
vin_chateau_lascombes]

erreur_appellation = ["APPELATIONS EN ERREUR : "]

# Normal avec lancement_initialisation_bd_blq.py
wb_bis : Workbook = openpyxl.load_workbook(ffl.dest_filename_test_bis)
# Lancer au niveau du fichier
# wb_bis : Workbook = openpyxl.load_workbook( "../" + ffl.dest_filename_production_bis )
# wb_bis : Workbook = openpyxl.load_workbook( "../" + ffl.dest_filename_test_bis )

def initialisation():

    print("Début de l'initialisation : " + str(datetime.now()))
    
    req.reinitialisation_global([nom.vin_str, nom.domaine_str])

    # wb_bis : Workbook = openpyxl.load_workbook(ffl.dest_filename_bis)
    # Lancer au niveau du fichier
    # wb_bis : Workbook = openpyxl.load_workbook( "../" + ffl.dest_filename_bis )
    # wb_bis = openpyxl.load_workbook( "../" + ffl.dest_filename_bis_test )

    print("Fin de l'initialisation : " + str(datetime.now()))

# Copier depuis nomenclature.pyeur 1
def reinitialisation_global(tab_table : list):
    
    for table in tab_table:

        requete = f"""DELETE FROM {table};"""

        req.envoie_requete_sans_retour(requete)

        requete = f"""ALTER TABLE {table} AUTO_INCREMENT = 0;"""

        req.envoie_requete_sans_retour(requete)

# Même fonction que traitement_applications : trouver un emplacement
def recupIdApresInsertion(tab_inverse : list, valeur_testee, nom_table : str, nom_colonne : str, id_colonne : int):
        
    id : int = req.recuperation_indice_doublon(nom_table, nom_colonne, valeur_testee)

    if(id != -1):
        return id

    # Mettre cette requête dans req !!!!
    # La "," est importante car sinon cela ne marche pas
    tuple_valeurs : tuple = (valeur_testee,)
    requete = f"""INSERT INTO """+ nom_table +""" ("""+ nom_colonne +""") VALUES (%s);"""

    req.envoie_requete_tuple_sans_retour(requete, tuple_valeurs)

    id : int = pp.recup_val_tab_inv2(tab_inverse, id_colonne, valeur_testee)

    return id

def parcourirDictionnaireRetourneId(valeur, dictionnaire):
    for valeur_dictionnaire in dictionnaire:
        if(valeur == valeur_dictionnaire):
            return dictionnaire[valeur_dictionnaire]
    return 1

def insertionDomaines():

    print("domaine")
    # Lecture du workbook d'origine
    sheet_list = wb_bis.sheetnames  # on récupère le nom des onglets, pas necessaire ici.
    sheet = wb_bis[sheet_list[0]]

    iDomaine = sheet['C']

    recupIdApresInsertion(tab_domaine_inverse, "NA", nom.domaine_str, "nom", 1)

    for domaine_valeur in liste_domaines_rajoutes:
    
        domaine : str = domaine_valeur

        portionCodeDoubleDomaine( domaine )

    for cellule_domaine in iDomaine :

        domaine = cellule_domaine.value

        portionCodeDoubleDomaine( domaine )


def portionCodeDoubleDomaine( domaine ):

    domaine = domaine.replace("'", " ")

    if( (len(domaine) == 0) or (domaine == "NA") or (domaine == "CHATEAU CHATEAU")):
        return

    # Domaines
    domaine_id = recupIdApresInsertion(tab_domaine_inverse, domaine, nom.domaine_str, "nom", 1)

# liste_appellation_a_enlever : list = ["CHATEAU D ISSAN", "MOULIS EN MEDOC", "CASTILLON-COTES DE BORDEAUX", "BORDEAUX ROUGE", "LALANDE DE POMEROL", "BLAYE-COTES DE BORDEAUX", "COTES DE BOURG"]

def fct_appellation_id(appellation : str):

    try:

        # print("appellation : " + appellation)

        # if( appellation in liste_appellation_a_enlever ):
        #     appellation = "MARGAUX"

        appellation_id = req.recuperation_un_id(nom.appellation_str, "appellation", appellation)

        # print("appellation_id : " + str(appellation_id))

        if( appellation_id == -1 ):

            erreur_appellation.append("Appellation en erreur : " + appellation)
            appellation_id = 1

        return appellation_id

    except ValueError:
        
        erreur_appellation.append(appellation + "")

    return 1

def insertion( vin, domaine_id, appellation_id, couleur_id, ws, index_sortie, couleur, iDomaine, i, region, iPays ):

    tuple_valeurs : tuple = (vin, "regex", domaine_id, appellation_id, couleur_id)

    requete = f"""INSERT INTO """ + nom.vin_str + """(nom, regex, domaine_id, appellation_id, couleur_id) VALUES (%s, %s, %s, %s, %s);"""
    req.envoie_requete_tuple_sans_retour(requete, tuple_valeurs) 

    # Fonction Attribution des valeurs dans les bonnes colonnes pour le nouveau workbook
    c1 = ws.cell(row = index_sortie, column = 1)
    c1.value = vin

    c2 = ws.cell(row = index_sortie, column = 2)
    c2.value = couleur

    c3 = ws.cell(row = index_sortie, column = 3)
    c3.value = iDomaine[i].value

    c4 = ws.cell(row = index_sortie, column = 4)
    c4.value = region

    c5 = ws.cell(row = index_sortie, column = 5)
    c5.value = iPays[i].value

def test1( liste_designation, tuple_valeurs, liste_valeurs ):

    [ vin, regex, domaine_id, appellation_id, couleur_id ] = liste_valeurs

    retour : bool = False

    nombre_doublon : int = req.compte_nombre_doublon( liste_designation, tuple_valeurs )

    # print("nombre_doublon : " + str(nombre_doublon))

    if( nombre_doublon > 0 ):

        print("Il y a déjà un vin pareil")

        if( nombre_doublon > 1 ):

            # supprimer les doublons
            requete = f"""DELETE FROM vin 
                WHERE nom = {vin} AND regex = {regex} AND domaine_id = {domaine_id} AND appellation_id = {appellation_id} AND couleur_id = {couleur_id} ;"""
            # print("requete : " + requete)

            req.envoie_requete_sans_retour(requete)

            return retour
        return True
    return retour

def traitement_vins_rajoutes(ws, index_sortie, iDomaine, region, iPays):

    for i  in range (0, len(liste_vins_rajoutes)):
            
        vin : str = liste_vins_rajoutes[i][0]

        retour_vin : str = ft.recupValeursListeUnD( liste_vin_a_tester, [vin] )

        if( retour_vin != "NA"):
            vin = retour_vin

        print("vin : " + vin)

        domaine = liste_vins_rajoutes[i][2]

        domaine_id : int = req.recuperation_un_id(nom.domaine_str, "nom", domaine)

        if( domaine_id == -1 ):
            domaine = ft.recupValeursListeUnD( liste_domaine_a_tester, [domaine] )

            domaine_id : int = req.recuperation_un_id(nom.domaine_str, "nom", domaine)

        # Appellation
        appellation : str = liste_vins_rajoutes[i][3]

        appellation_id = fct_appellation_id(appellation)

        # print("appellation_id : " + str(appellation_id))

        # Couleur
        couleur : str = liste_vins_rajoutes[i][4]

        couleur = ft.recupValeursListeDeuxD( liste_couleur, [couleur] )

        couleur_id : int = req.recuperation_un_id(nom.couleur_vin_str, "couleur", couleur)

        ###
        liste_designation : list = ["vin", "nom", "regex", "domaine_id", "appellation_id", "couleur_id"]
        tuple_valeurs : tuple = ( vin, "regex", domaine_id, appellation_id, couleur_id )

        liste_valeurs = [ vin, "regex", domaine_id, appellation_id, couleur_id ]

        if( test1( liste_designation, tuple_valeurs, liste_valeurs ) ):
            continue
        
        insertion( vin, domaine_id, appellation_id, couleur_id, ws, index_sortie, couleur, iDomaine, i, region, iPays )

        index_sortie += 1
    return index_sortie


def traitement_vins_LWIN(index_sortie, row_count, ws, iVinAppellation, iRegion, iDomaine, iCouleur, iPays, iAppellation ):

    for i  in range (1, row_count):
            
        vin : str = iVinAppellation[i].value
        print("vin")
        print(vin)

        liste_vinApellation  = vin.split(',')

        print(liste_vinApellation)

        vin = vin.replace(",", '')

        if( len(vin) == 0 ):
            vin = "NA"
            return

        region = (iRegion[i].value).upper()

        domaine : list = (iDomaine[i].value)

        # print("tab_vinApellation[0] : " + list_vinApellation[0])

        # !!! A mettre en place !!!
        # Revoir si exception, les mettre à la main manuellement
        # Je ne sais pas d'où vient ce test et ce qu'il fait
        if( len( liste_vinApellation ) == 1):
            print( " list_vinApellation == 1 " )

        elif( region == "BORDEAUX"):
          vin = liste_vinApellation[0]

          print("Passe bordeaux")
        
        elif( liste_vinApellation[0] == domaine ):
    
            vin : str = ''.join(liste_vinApellation[1:]).strip() 

        else:
          print("Erreur")
          continue

        print("vin !!! : " + vin)

        couleur : str = iCouleur[i].value

        # print("couleur : ")
        # print( couleur )

        # Couleur
        couleur = ft.recupValeursListeDeuxD( liste_couleur, [couleur] )

        # print("couleur : " + couleur)

        couleur_id : int = req.recuperation_un_id(nom.couleur_vin_str, "couleur", couleur)

        # print("couleur_id : " + str(couleur_id))

        pays : int = iPays[i].value

        pays_id : int = req.recuperation_un_id(nom.pays_str, "nom", pays)

        # print("pays_id  : -" + str(pays_id) + "-")

        # Insérer les lignes de domaines

        # Domaine
        # print("domaine  : -" + domaine + "-")

        domaine_id : int = req.recuperation_un_id(nom.domaine_str, "nom", domaine)

        # print("domaine_id : " + str(domaine_id))

        if( domaine_id == -1 ):
            domaine = ft.recupValeursListeUnD( liste_domaine_a_tester, [domaine] )

            domaine_id : int = req.recuperation_un_id(nom.domaine_str, "nom", domaine)

        # Appellation
        # appellation : str = "BORDEAUX"
        appellation : str = iAppellation[i].value

        # if( len(vinAppellation) == 2 ):
        #     appellation = vinAppellation[1].strip()

        appellation_id = fct_appellation_id(appellation)

        print("appellation_id : " + str(appellation_id))

        print("vin : " + vin + " regex : " + str(domaine_id) + " appellation_id : " + str(appellation_id) + " couleur_id : " + str(couleur_id))

        liste_designation : list = ["vin", "nom", "regex", "domaine_id", "appellation_id", "couleur_id"]
        tuple_valeurs : tuple = ( vin, "regex", domaine_id, appellation_id, couleur_id )

        liste_valeurs = [ vin, "regex", domaine_id, appellation_id, couleur_id ]

        if( test1( liste_designation, tuple_valeurs, liste_valeurs ) ):
            continue
        
        insertion( vin, domaine_id, appellation_id, couleur_id, ws, index_sortie, couleur, iDomaine, i, region, iPays )

        index_sortie += 1

    return index_sortie


def insertionVins():

    print("vin")

    req.reinitialisation_global([nom.vin_str])

    # le rajouter dans req !!!!
    # Insertion d'une ligne, pour les vins non identifiés
    tuple_valeurs : tuple = ("NA", "NA", 1, 1, 1)
    requete = f"""INSERT INTO """ + nom.vin_str + """(nom, regex, domaine_id, appellation_id, couleur_id) VALUES (%s, %s, %s, %s, %s);"""
    req.envoie_requete_tuple_sans_retour(requete, tuple_valeurs) 

    # On prépare un nouveau workbook
    ###
    wb3 = Workbook()
    ws = wb3.active
    ws.title = ffl.nomProfil
    ###

    # Lecture du workbook d'origine
    sheet_list = wb_bis.sheetnames  # on récupère le nom des onglets, pas necessaire ici.
    sheet = wb_bis[sheet_list[0]]

    iVinAppellation = sheet['A']
    iCouleur = sheet['B']
    iDomaine = sheet['C']
    iRegion = sheet['D']
    region : str = ""
    iPays = sheet['E']

    iAppellation = sheet['F']

    row_count = sheet.max_row

    index_sortie = 1

    # index_sortie = traitement_vins_rajoutes(ws, index_sortie, iDomaine, region, iPays)

    index_sortie = traitement_vins_LWIN(index_sortie, row_count, ws, iVinAppellation, iRegion, iDomaine, iCouleur, iPays, iAppellation)
    
    # Ecriture du nouveau workbook : a modifier pour que l'on ai pas à commenter/décommenter tout le temps
    wb3.save(ffl.dest_filename_production_final)
    # wb3.save(ffl.dest_filename_test_final)

    # wb3.save( "/var/www/html/ressources/nomenclature/ressources_bd_blq/toto.csv ")
        
def recuperationVins():

    initialisation()

    try:

        print("Insertion des domaines : " + str(datetime.now()))

        insertionDomaines()

        # L'insertion des vins passe avant l'insertion des domaines, comme le vin a une clef secondaire de domaine affiche une erreur.
        # time.sleep(2)

        tab_domaine = req.recuperation_tab( nom.domaine_str )
        tab_domaine_inverse = list(map(list, np.transpose(tab_domaine)))

        print("Fin insertion des domaines : " + str(datetime.now()))

        print("Insertion des vins : " + str(datetime.now()))
        
        insertionVins()

        print("Fin insertion des vins : " + str(datetime.now()))

        # Les erreurs
        print("Les erreurs : ")
        print(erreur_appellation)

    except Exception as erreur :
        print(type(erreur))    # the exception instance
        print(erreur.args)     # arguments stored in .args
        print(erreur)

    # os.remove(dest_filename_bis)
    # os.remove(dest_filename)

# recuperationVins()








   

